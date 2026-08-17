"""Módulo de Reportes categorizados con exportación a Excel y PDF.

Categorías disponibles:
- deliveries_by_employee_date: Equipos entregados por empleado en un día específico
- returns_by_period: Devoluciones por período
- assets_by_status: Equipos por estado
- assets_by_department: Equipos por departamento
- assignment_history: Historial de asignaciones
"""
import io
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle)

from core.database import db
from core.deps import require_permission

router = APIRouter(prefix="/reports", tags=["Reportes"])


# ---------- Configuración de categorías ----------
CATEGORIES = {
    "deliveries_by_employee_date": {
        "title": "Equipos entregados por empleado en un día",
        "columns": [
            ("delivery_date", "Fecha"),
            ("asset_tag", "Etiqueta"),
            ("asset_name", "Equipo"),
            ("serial_number", "N° Serie"),
            ("mac_address", "MAC"),
            ("employee_code", "Cód. Empleado"),
            ("assigned_to_name", "Empleado"),
            ("department_name", "Departamento"),
            ("delivered_by", "Entregado Por"),
            ("condition", "Condición"),
        ],
    },
    "returns_by_period": {
        "title": "Devoluciones por período",
        "columns": [
            ("return_date", "Fecha Dev."),
            ("asset_tag", "Etiqueta"),
            ("asset_name", "Equipo"),
            ("serial_number", "N° Serie"),
            ("returned_by", "Devuelto Por"),
            ("employee_code", "Cód. Empleado"),
            ("received_by", "Recibido Por"),
            ("condition", "Condición"),
        ],
    },
    "assets_by_status": {
        "title": "Equipos por estado",
        "columns": [
            ("asset_tag", "Etiqueta"),
            ("name", "Nombre"),
            ("serial_number", "N° Serie"),
            ("category_name", "Categoría"),
            ("status", "Estado"),
            ("assigned_to_name", "Responsable"),
            ("department_name", "Departamento"),
            ("branch_name", "Sucursal"),
        ],
    },
    "assets_by_department": {
        "title": "Equipos por departamento",
        "columns": [
            ("asset_tag", "Etiqueta"),
            ("name", "Nombre"),
            ("serial_number", "N° Serie"),
            ("category_name", "Categoría"),
            ("status", "Estado"),
            ("assigned_to_name", "Responsable"),
            ("department_name", "Departamento"),
        ],
    },
    "assignment_history": {
        "title": "Historial de asignaciones",
        "columns": [
            ("delivery_date", "Fecha Entrega"),
            ("asset_tag", "Etiqueta"),
            ("asset_name", "Equipo"),
            ("serial_number", "N° Serie"),
            ("employee_code", "Cód. Empleado"),
            ("assigned_to_name", "Empleado"),
            ("department_name", "Departamento"),
            ("delivered_by", "Entregado Por"),
            ("returned", "Devuelto"),
            ("return_date", "Fecha Devolución"),
        ],
    },
}


def _fmt_date(v):
    if not v:
        return ""
    try:
        return datetime.fromisoformat(v.replace("Z", "+00:00")).strftime("%d/%m/%Y")
    except Exception:
        return str(v)[:10]


def _fmt_datetime(v):
    if not v:
        return ""
    try:
        return datetime.fromisoformat(v.replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(v)


def _fmt_bool(v):
    return "Sí" if v else "No"


async def _resolve_names(items: List[dict], keys_lookup: Dict[str, str]) -> None:
    """Enriquece items con nombres de referencias (categorías, deptos, sucursales)."""
    for out_key, coll in keys_lookup.items():
        src_key = out_key.replace("_name", "_id")
        ids = list({i.get(src_key) for i in items if i.get(src_key)})
        if not ids:
            continue
        docs = await db[coll].find({"id": {"$in": ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
        m = {d["id"]: d.get("name") for d in docs}
        for it in items:
            it[out_key] = m.get(it.get(src_key))


async def _fetch_data(category: str, filters: dict) -> List[dict]:
    if category == "deliveries_by_employee_date":
        emp_id = filters.get("employee_id")
        date = filters.get("date")
        if not date:
            raise HTTPException(status_code=400, detail="Debe indicar la fecha (date)")
        q = {"deleted": {"$ne": True},
             "delivery_date": {"$gte": f"{date}T00:00:00",
                                "$lt":  f"{date}T23:59:59.999999\uffff"}}
        if emp_id:
            q["employee_id"] = emp_id
        elif filters.get("employee_code"):
            q["employee_code"] = filters["employee_code"]
        items = await db.deliveries.find(q, {"_id": 0}).sort("delivery_date", -1).to_list(2000)
        await _resolve_names(items, {"department_name": "departments"})
        return items

    if category == "returns_by_period":
        f = filters.get("from_date")
        t = filters.get("to_date")
        q = {}
        rng = {}
        if f:
            rng["$gte"] = f
        if t:
            rng["$lte"] = f"{t}T23:59:59.999999"
        if rng:
            q["return_date"] = rng
        items = await db.returns.find(q, {"_id": 0}).sort("return_date", -1).to_list(5000)
        return items

    if category == "assets_by_status":
        status = filters.get("status")
        q = {"deleted": {"$ne": True}}
        if status and status != "todos":
            q["status"] = status
        items = await db.assets.find(q, {"_id": 0}).sort("asset_tag", 1).to_list(10000)
        await _resolve_names(items, {"category_name": "categories",
                                     "department_name": "departments",
                                     "branch_name": "branches"})
        return items

    if category == "assets_by_department":
        dep = filters.get("department_id")
        q = {"deleted": {"$ne": True}}
        if dep and dep != "todos":
            q["department_id"] = dep
        items = await db.assets.find(q, {"_id": 0}).sort("department_id", 1).to_list(10000)
        await _resolve_names(items, {"category_name": "categories",
                                     "department_name": "departments"})
        return items

    if category == "assignment_history":
        emp_id = filters.get("employee_id")
        f = filters.get("from_date")
        t = filters.get("to_date")
        q = {"deleted": {"$ne": True}}
        if emp_id:
            q["employee_id"] = emp_id
        rng = {}
        if f:
            rng["$gte"] = f
        if t:
            rng["$lte"] = f"{t}T23:59:59.999999"
        if rng:
            q["delivery_date"] = rng
        items = await db.deliveries.find(q, {"_id": 0}).sort("delivery_date", -1).to_list(10000)
        await _resolve_names(items, {"department_name": "departments"})
        return items

    raise HTTPException(status_code=400, detail=f"Categoría de reporte desconocida: {category}")


def _row_values(item: dict, columns: List[tuple]) -> List[Any]:
    out = []
    for key, _label in columns:
        v = item.get(key)
        if key in ("delivery_date", "return_date"):
            v = _fmt_datetime(v)
        elif key == "returned":
            v = _fmt_bool(v)
        elif v is None:
            v = ""
        out.append(v)
    return out


@router.get("/categories")
async def list_categories(user: dict = Depends(require_permission("reports:read"))):
    return {"items": [{"key": k, "title": v["title"]} for k, v in CATEGORIES.items()]}


@router.get("/preview")
async def preview(category: str,
                  employee_id: Optional[str] = None,
                  employee_code: Optional[str] = None,
                  date: Optional[str] = None,
                  from_date: Optional[str] = None,
                  to_date: Optional[str] = None,
                  status: Optional[str] = None,
                  department_id: Optional[str] = None,
                  user: dict = Depends(require_permission("reports:read"))):
    if category not in CATEGORIES:
        raise HTTPException(status_code=400, detail="Categoría inválida")
    filters = dict(employee_id=employee_id, employee_code=employee_code, date=date,
                   from_date=from_date, to_date=to_date, status=status,
                   department_id=department_id)
    data = await _fetch_data(category, filters)
    cfg = CATEGORIES[category]
    rows = [_row_values(it, cfg["columns"]) for it in data]
    return {"title": cfg["title"],
            "columns": [{"key": k, "label": lbl} for k, lbl in cfg["columns"]],
            "rows": rows, "total": len(rows), "filters": filters}


def _excel_bytes(title: str, columns: List[tuple], data: List[dict],
                 filters: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="CC0000")
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value=f"Cisa TI · {title}").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = center

    # metadata
    meta = [f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}"]
    for k, v in filters.items():
        if v:
            meta.append(f"{k}: {v}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws.cell(row=2, column=1, value=" | ".join(meta)).font = Font(italic=True, size=9,
                                                                    color="666666")

    # headers
    hdr_row = 4
    for c, (_key, label) in enumerate(columns, start=1):
        cell = ws.cell(row=hdr_row, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # rows
    for r, item in enumerate(data, start=hdr_row + 1):
        for c, (key, _label) in enumerate(columns, start=1):
            v = item.get(key)
            if key in ("delivery_date", "return_date"):
                v = _fmt_datetime(v)
            elif key == "returned":
                v = _fmt_bool(v)
            ws.cell(row=r, column=c, value=v)

    # ancho auto
    for c, (_key, label) in enumerate(columns, start=1):
        max_len = max(len(str(label)), 8)
        for r in range(hdr_row + 1, hdr_row + 1 + len(data)):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                max_len = max(max_len, min(50, len(str(val))))
        ws.column_dimensions[get_column_letter(c)].width = max_len + 2

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_bytes(title: str, columns: List[tuple], data: List[dict],
               filters: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            title=title)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(f"<b>Cisa TI</b> · {title}", styles["Title"]))
    meta_parts = [f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}"]
    for k, v in filters.items():
        if v:
            meta_parts.append(f"{k}: {v}")
    story.append(Paragraph(" | ".join(meta_parts), styles["Italic"]))
    story.append(Spacer(1, 4 * mm))

    header = [lbl for _k, lbl in columns]
    rows = [header]
    for it in data:
        rows.append([str(x) if x is not None else "" for x in _row_values(it, columns)])

    if len(rows) == 1:
        story.append(Paragraph("Sin datos para el filtro aplicado.", styles["Normal"]))
    else:
        table = Table(rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#CC0000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c9d1d9")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f7f8fa")]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(table)
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(f"Total de registros: <b>{len(data)}</b>", styles["Normal"]))
    doc.build(story)
    return buf.getvalue()


@router.get("/export")
async def export_report(category: str,
                        format: str = Query(..., pattern="^(xlsx|pdf)$"),
                        employee_id: Optional[str] = None,
                        employee_code: Optional[str] = None,
                        date: Optional[str] = None,
                        from_date: Optional[str] = None,
                        to_date: Optional[str] = None,
                        status: Optional[str] = None,
                        department_id: Optional[str] = None,
                        user: dict = Depends(require_permission("reports:read"))):
    if category not in CATEGORIES:
        raise HTTPException(status_code=400, detail="Categoría inválida")
    filters = dict(employee_id=employee_id, employee_code=employee_code, date=date,
                   from_date=from_date, to_date=to_date, status=status,
                   department_id=department_id)
    data = await _fetch_data(category, filters)
    cfg = CATEGORIES[category]
    fname_base = f"{category}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if format == "xlsx":
        content = _excel_bytes(cfg["title"], cfg["columns"], data, filters)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname_base}.xlsx"'})

    content = _pdf_bytes(cfg["title"], cfg["columns"], data, filters)
    return StreamingResponse(
        io.BytesIO(content), media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname_base}.pdf"'})
